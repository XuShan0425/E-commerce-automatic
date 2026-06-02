"""店铺商品与数据分析导入 API — 从速卖通导出文件智能导入."""

from __future__ import annotations

import os
import re
import time
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from App.core.database import get_db
from App.core.logging import get_logger
from App.core.security import verify_api_key
from App.models.base import Product
from App.models.product_analytics import (
    CoreMetric,
    KeywordData,
    ServiceData,
    SkuAnalysis,
    TrafficSource,
)
from App.services.cookie_manager import CookieManager

logger = get_logger(__name__)

router = APIRouter(prefix="/import", tags=["import"])

_browser_state: dict[str, Any] = {
    "pid": None, "started_at": None, "service": None,
}

# ── 常量 ────────────────────────────────────────

_DOWNLOADS_DIR = Path(os.environ.get("USERPROFILE", "C:/Users/Default")) / "Downloads"

_CSP_PRODUCT_LIST_URL = (
    "https://csp.aliexpress.com/supplier/product/list?source=page_center"
)

# SYCM 文件名模式 → 数据类型映射
_SYCM_FILE_PATTERNS: list[tuple[str, str, str]] = [
    (r"生意参谋.*来源明细", "traffic_source", "流量来源明细"),
    (r"单品分析.*核心指标.*国家分析", "core_metric_country", "核心指标(分国家)"),
    (r"单品分析.*核心指标", "core_metric", "核心指标"),
    (r"单品分析.*SKU", "sku_analysis", "SKU分析"),
    (r"单品分析.*关键词", "keyword_data", "关键词数据"),
    (r"服务分析.*明细数据", "service_data", "服务分析数据"),
    (r"商品明细.*来源明细", "traffic_source", "流量来源明细"),
]

# ── 文件检测 ─────────────────────────────────────

def _detect_file_type(filename: str) -> tuple[str, str]:
    """检测文件类型，返回 (type_code, type_label)。

    优先匹配 SYCM 分析数据，回退为商品列表。
    """
    for pattern, type_code, type_label in _SYCM_FILE_PATTERNS:
        if re.search(pattern, filename):
            return type_code, type_label
    return "product_list", "商品列表"


def _find_downloaded_xlsx_files() -> list[dict]:
    """扫描 Downloads 文件夹中最近1小时的 XLSX 文件。"""
    if not _DOWNLOADS_DIR.exists():
        return []

    cutoff = time.time() - 3600
    files: list[dict] = []

    for f in _DOWNLOADS_DIR.iterdir():
        if not f.is_file():
            continue
        ext = f.suffix.lower()
        if ext not in (".xlsx", ".xls"):
            continue
        mtime = f.stat().st_mtime
        if mtime < cutoff:
            continue

        type_code, type_label = _detect_file_type(f.name)
        size_kb = round(f.stat().st_size / 1024, 1)
        files.append({
            "name": f.name,
            "path": str(f),
            "size_kb": size_kb,
            "modified": datetime.fromtimestamp(mtime, tz=timezone.utc).isoformat(),
            "file_type": type_code,
            "file_type_label": type_label,
        })

    files.sort(key=lambda x: x["modified"], reverse=True)
    return files


# ── XLSX 解析 ────────────────────────────────────

def _read_xlsx_headers_and_rows(file_path: str) -> dict[str, Any]:
    """读取 XLSX 所有 sheet 的表头和数据行。"""
    import openpyxl
    wb = openpyxl.load_workbook(file_path, data_only=True)
    result: dict[str, Any] = {"sheets": {}}
    total = 0

    for sname in wb.sheetnames:
        if "_hide" in sname or "隐藏" in sname:
            continue
        ws = wb[sname]
        if ws.max_row < 1:
            continue

        # 读取第1行作为表头
        headers: list[str] = []
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row=1, column=col).value
            headers.append(str(v).strip() if v else f"col_{col}")

        data_rows: list[list[str]] = []
        for ri in range(2, ws.max_row + 1):
            vals: list[str] = []
            has_any = False
            for col in range(1, ws.max_column + 1):
                v = ws.cell(row=ri, column=col).value
                s = ""
                if v is not None:
                    if isinstance(v, float):
                        s = str(int(v)) if v == int(v) and v < 1e18 else str(v)
                    else:
                        s = str(v).strip()
                    if s:
                        has_any = True
                vals.append(s)
            if has_any:
                data_rows.append(vals)
                total += 1

        result["sheets"][sname] = {
            "headers": headers,
            "rows": data_rows,
            "count": len(data_rows),
        }

    wb.close()
    result["total_rows"] = total
    return result


# ── 各类型解析器 ──────────────────────────────────

def _parse_product_list(file_path: str) -> list[dict]:
    """解析商品列表导出文件（CSP 导出格式）。"""
    import openpyxl
    wb = openpyxl.load_workbook(file_path, data_only=True)

    ALL_ALIASES = {
        "id": "sku_id", "sku": "sku_id",
        "*商品标题": "name", "商品标题": "name", "商品名称": "name",
    }
    seen_sku = set()
    all_rows: list[dict] = []

    for sname in wb.sheetnames:
        if "_hide" in sname:
            continue
        ws = wb[sname]
        if ws.max_row < 2:
            continue
        # 表头在第2行（第1行是分类标题）
        headers: list[str] = []
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row=2, column=col).value
            headers.append(str(v).strip() if v else "")

        field_map: dict[int, str] = {}
        for ci, raw in enumerate(headers):
            key = raw.lower().replace(" ", "_").replace("-", "_")
            key = ALL_ALIASES.get(key, key)
            field_map[ci] = key

        for ri in range(3, ws.max_row + 1):
            vals: list[str] = []
            has_any = False
            for col in range(1, ws.max_column + 1):
                v = ws.cell(row=ri, column=col).value
                s = ""
                if v is not None:
                    if isinstance(v, float):
                        s = str(int(v)) if v == int(v) else str(v)
                    else:
                        s = str(v).strip()
                    if s:
                        has_any = True
                vals.append(s)
            if not has_any:
                continue

            rd: dict[str, str] = {}
            for ci, key in field_map.items():
                val = vals[ci] if ci < len(vals) else ""
                if key in rd:
                    if val and not rd[key]:
                        rd[key] = val
                else:
                    rd[key] = val

            sku_id = rd.get("sku_id", "")
            if not sku_id or sku_id in seen_sku:
                continue
            seen_sku.add(sku_id)
            name = re.sub(r'\s+', ' ', rd.get("name", "")).strip()
            all_rows.append({
                "sku_id": sku_id,
                "name": name[:80],
                "category": rd.get("category", sname),
            })

    wb.close()
    return all_rows


def _parse_float(val: str) -> float:
    try:
        return float(val.replace(",", "").replace("-", "0"))
    except (ValueError, AttributeError):
        return 0.0


def _parse_int(val: str) -> int:
    try:
        return int(val.replace(",", "").replace("-", "0"))
    except (ValueError, AttributeError):
        return 0


def _to_stat_date(val: str) -> date:
    """解析统计日期，兼容 20260529 和 2026-04-30~2026-05-29 格式。"""
    val = val.strip()
    if "~" in val:
        val = val.split("~")[-1]
    val = val.replace("-", "")
    try:
        return date(int(val[:4]), int(val[4:6]), int(val[6:8]))
    except (ValueError, IndexError):
        return date.today()


def _parse_core_metric(file_path: str) -> list[dict]:
    """解析核心指标文件 → CoreMetric。"""
    xls = _read_xlsx_headers_and_rows(file_path)
    results: list[dict] = []
    seen = set()

    for sname, sheet in xls["sheets"].items():
        headers = [h.lower().replace(" ", "_").replace("-", "_") for h in sheet["headers"]]
        for row in sheet["rows"]:
            data = dict(zip(headers, row))
            sku_id = data.get("商品id", data.get("product_id", ""))
            stat_date = _to_stat_date(data.get("统计日期", ""))
            if not sku_id:
                continue
            key = f"{sku_id}_{stat_date}"
            if key in seen:
                continue
            seen.add(key)

            metrics = {}
            metric_fields = {
                "搜索曝光量": "search_impressions",
                "搜索点击率": "search_ctr",
                "商品访客数": "visitors",
                "商品浏览量": "page_views",
                "平均停留时长": "avg_stay_seconds",
                "商品加购人数": "cart_add_users",
                "商品收藏人数": "favorites",
                "下单金额": "order_amount",
                "下单买家数": "order_buyers",
                "下单转化率": "order_conversion",
                "支付金额": "payment_amount",
                "支付买家数": "payment_buyers",
                "支付转化率": "payment_conversion",
                "访客数": "visitors",
                "跳失率": "bounce_rate",
            }
            for cn, en in metric_fields.items():
                if cn in data:
                    metrics[en] = _parse_float(data[cn])

            results.append({
                "sku_id": sku_id,
                "stat_date": stat_date.isoformat(),
                "metrics": metrics,
                "country": data.get("国家id", data.get("国家", "")),
            })

    return results


def _parse_keyword_data(file_path: str) -> list[dict]:
    """解析关键词文件 → KeywordData。"""
    xls = _read_xlsx_headers_and_rows(file_path)
    results: list[dict] = []

    for sheet in xls["sheets"].values():
        headers = [h.lower().replace(" ", "_").replace("-", "_") for h in sheet["headers"]]
        for row in sheet["rows"]:
            data = dict(zip(headers, row))
            sku_id = data.get("商品id", "")
            keyword = data.get("关键词", "")
            stat_date = _to_stat_date(data.get("统计日期", ""))
            if not sku_id or not keyword:
                continue

            metrics = {}
            kw_fields = {
                "搜索曝光量": "search_impressions",
                "搜索曝光人数": "search_exposure_users",
                "词引导浏览量": "keyword_page_views",
                "词引导访客数": "keyword_visitors",
                "词引导支付金额": "keyword_payment",
                "词引导支付买家数": "keyword_payment_buyers",
                "词引导支付转化率": "keyword_payment_conversion",
                "词引导支付订单数": "keyword_payment_orders",
            }
            for cn, en in kw_fields.items():
                if cn in data:
                    metrics[en] = _parse_float(data[cn])

            results.append({
                "sku_id": sku_id,
                "stat_date": stat_date.isoformat(),
                "keyword": keyword,
                "metrics": metrics,
            })

    return results


def _parse_sku_analysis(file_path: str) -> list[dict]:
    """解析 SKU 分析文件 → SkuAnalysis。"""
    xls = _read_xlsx_headers_and_rows(file_path)
    results: list[dict] = []

    for sheet in xls["sheets"].values():
        headers = [h.lower().replace(" ", "_").replace("-", "_") for h in sheet["headers"]]
        for row in sheet["rows"]:
            data = dict(zip(headers, row))
            sku_id = data.get("商品id", "")
            sku_code = data.get("sku_id", data.get("skuid", ""))
            stat_date = _to_stat_date(data.get("统计日期", ""))
            if not sku_id:
                continue

            metrics = {}
            sku_fields = {
                "sku支付金额": "payment_amount",
                "sku支付买家数": "payment_buyers",
                "sku支付件数": "payment_quantity",
                "sku加购买家数": "cart_add_buyers",
                "sku加购件数": "cart_add_quantity",
            }
            for cn, en in sku_fields.items():
                if cn in data:
                    metrics[en] = _parse_float(data[cn])

            results.append({
                "sku_id": sku_id,
                "stat_date": stat_date.isoformat(),
                "sku_code": sku_code,
                "sku_info": data.get("sku信息", ""),
                "metrics": metrics,
            })

    return results


def _parse_traffic_source(file_path: str) -> list[dict]:
    """解析流量来源文件 → TrafficSource。"""
    xls = _read_xlsx_headers_and_rows(file_path)
    results: list[dict] = []

    for sheet in xls["sheets"].values():
        headers = [h.lower().replace(" ", "_").replace("-", "_") for h in sheet["headers"]]
        for row in sheet["rows"]:
            data = dict(zip(headers, row))
            source = data.get("流量来源", "")
            sub_source = data.get("二级来源", "")
            stat_date = _to_stat_date(data.get("统计日期", ""))
            if not source:
                continue

            metrics = {}
            tf_fields = {
                "访客数": "visitors",
                "访客数占比": "visitor_ratio",
                "商品加购人数": "cart_add_users",
                "商品收藏人数": "favorites",
                "平均访问深度": "avg_depth",
                "跳失率": "bounce_rate",
                "下单金额": "order_amount",
                "下单买家数": "order_buyers",
                "下单转化率": "order_conversion",
                "支付金额": "payment_amount",
                "支付买家数": "payment_buyers",
                "支付转化率": "payment_conversion",
                "uv价值": "uv_value",
            }
            for cn, en in tf_fields.items():
                if cn in data:
                    metrics[en] = _parse_float(data[cn])

            results.append({
                "stat_date": stat_date.isoformat(),
                "source_name": source,
                "sub_source": sub_source,
                "metrics": metrics,
            })

    return results


def _parse_service_data(file_path: str) -> list[dict]:
    """解析服务分析文件 → ServiceData。"""
    xls = _read_xlsx_headers_and_rows(file_path)
    results: list[dict] = []

    for sheet in xls["sheets"].values():
        headers = [h.lower().replace(" ", "_").replace("-", "_") for h in sheet["headers"]]
        for row in sheet["rows"]:
            data = dict(zip(headers, row))
            sku_id = data.get("产品id", data.get("商品id", ""))
            stat_date_str = data.get("不良体验发生时间") or data.get("统计日期") or ""
            stat_date = _to_stat_date(stat_date_str)
            if not sku_id:
                continue

            results.append({
                "sku_id": sku_id,
                "stat_date": stat_date.isoformat(),
                "metrics": {k: v for k, v in data.items() if v and k not in ("产品id", "商品id", "统计日期", "不良体验发生时间")},
            })

    return results


_PARSERS: dict[str, Any] = {
    "product_list": _parse_product_list,
    "traffic_source": _parse_traffic_source,
    "core_metric": _parse_core_metric,
    "core_metric_country": _parse_core_metric,
    "sku_analysis": _parse_sku_analysis,
    "keyword_data": _parse_keyword_data,
    "service_data": _parse_service_data,
}

_FILE_TYPE_LABELS: dict[str, str] = {
    "product_list": "商品列表",
    "traffic_source": "流量来源",
    "core_metric": "核心指标",
    "core_metric_country": "核心指标(分国家)",
    "sku_analysis": "SKU分析",
    "keyword_data": "关键词",
    "service_data": "服务分析",
}


# ── API 端点 ─────────────────────────────────────

@router.post("/open-export-page")
async def open_export_page(
    _api_key: str = Depends(verify_api_key),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """打开速卖通后台商品管理页，供用户手动导出。"""
    global _browser_state

    if _browser_state.get("pid"):
        elapsed = time.time() - (_browser_state.get("started_at") or 0)
        if elapsed < 300:
            return {"status": "ok", "message": "浏览器已在运行中", "elapsed_seconds": round(elapsed)}
        _browser_state = {"pid": None, "started_at": None, "service": None}

    cookie_mgr = CookieManager(db)
    cookies = await cookie_mgr.load_cookies("aliexpress.com")
    if not cookies:
        raise HTTPException(status_code=400, detail="未找到速卖通登录信息，请先在系统设置中完成登录")

    def _open_browser(cookies: list[dict]) -> None:
        from App.services.browser import BrowserService
        bs = BrowserService(headless=False)
        context = bs.new_context(cookies=cookies)
        page = context.new_page()
        page.goto(_CSP_PRODUCT_LIST_URL, wait_until="domcontentloaded", timeout=30_000)
        page.wait_for_timeout(5_000)
        if "login" in page.url.lower():
            logger.warning("Cookie 已过期，浏览器已打开但需要重新登录")
        else:
            logger.info("已打开商品管理页: %s", page.url[:80])
        _browser_state["service"] = bs
        import threading
        stop_event = threading.Event()
        _browser_state["_stop_event"] = stop_event
        stop_event.wait()

    import threading
    thread = threading.Thread(target=_open_browser, args=(cookies,), daemon=True)
    thread.start()
    time.sleep(3)
    _browser_state.update({"pid": os.getpid(), "started_at": time.time()})
    return {"status": "ok", "message": "速卖通后台已打开，请在浏览器中导出数据"}


@router.get("/check-export-files")
async def check_export_files(
    _api_key: str = Depends(verify_api_key),
) -> dict:
    """扫描下载文件夹中最近1小时的 XLSX 文件，自动识别类型。"""
    files = _find_downloaded_xlsx_files()
    return {"status": "ok", "files": files, "count": len(files)}


@router.post("/preview-export-file")
async def preview_export_file(
    filename: str,
    _api_key: str = Depends(verify_api_key),
) -> dict:
    """预览指定文件，按类型解析后返回结构化数据。"""
    files = _find_downloaded_xlsx_files()
    matched = [f for f in files if f["name"] == filename]
    if not matched:
        raise HTTPException(status_code=404, detail=f"文件 '{filename}' 未找到或已过期")

    file_path = matched[0]["path"]
    file_type = matched[0]["file_type"]
    parser = _PARSERS.get(file_type)
    if not parser:
        raise HTTPException(status_code=400, detail=f"不支持的文件类型: {file_type}")

    try:
        data = parser(file_path)
    except Exception as exc:
        logger.exception("解析失败")
        raise HTTPException(status_code=400, detail=f"解析文件失败: {exc}")

    return {
        "status": "ok",
        "filename": filename,
        "file_type": file_type,
        "file_type_label": _FILE_TYPE_LABELS.get(file_type, file_type),
        "data": data[:50],  # 预览前50条
        "count": len(data),
    }


@router.post("/import-analytics")
async def import_analytics(
    filename: str,
    _api_key: str = Depends(verify_api_key),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """将 SYCM 分析数据导入数据库。"""
    files = _find_downloaded_xlsx_files()
    matched = [f for f in files if f["name"] == filename]
    if not matched:
        raise HTTPException(status_code=404, detail=f"文件 '{filename}' 未找到或已过期")

    file_path = matched[0]["path"]
    file_type = matched[0]["file_type"]

    if file_type == "product_list":
        raise HTTPException(status_code=400, detail="请使用商品导入接口导入此文件")

    parser = _PARSERS.get(file_type)
    if not parser:
        raise HTTPException(status_code=400, detail=f"不支持的文件类型: {file_type}")

    try:
        records = parser(file_path)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"解析文件失败: {exc}")

    if not records:
        raise HTTPException(status_code=400, detail="文件中没有有效数据")

    # 批量查询已有 SKU 映射
    all_sku_ids = list(set(r.get("sku_id", "") for r in records if r.get("sku_id")))
    sku_map: dict[str, int] = {}
    if all_sku_ids:
        result = await db.execute(select(Product.id, Product.sku_id).where(Product.sku_id.in_(all_sku_ids)))
        for row in result.all():
            sku_map[row.sku_id] = row.id

    imported = 0
    skipped = 0

    if file_type == "traffic_source":
        for r in records:
            _create_traffic_source(db, r)
            imported += 1

    elif file_type in ("core_metric", "core_metric_country"):
        seen_keys = set()
        for r in records:
            product_id = sku_map.get(r["sku_id"])
            if not product_id:
                skipped += 1
                continue
            key = f"{r['sku_id']}_{r['stat_date']}_{r.get('country', '')}"
            if key in seen_keys:
                skipped += 1
                continue
            seen_keys.add(key)
            stat_date = date.fromisoformat(r["stat_date"])
            db.add(CoreMetric(
                product_id=product_id,
                stat_date=stat_date,
                metrics={
                    **r["metrics"],
                    **({"country": r["country"]} if r.get("country") else {}),
                },
            ))
            imported += 1

    elif file_type == "keyword_data":
        for r in records:
            product_id = sku_map.get(r["sku_id"])
            if not product_id:
                skipped += 1
                continue
            stat_date = date.fromisoformat(r["stat_date"])
            db.add(KeywordData(
                product_id=product_id,
                stat_date=stat_date,
                keyword=r["keyword"],
                metrics=r["metrics"],
            ))
            imported += 1

    elif file_type == "sku_analysis":
        for r in records:
            product_id = sku_map.get(r["sku_id"])
            if not product_id:
                skipped += 1
                continue
            stat_date = date.fromisoformat(r["stat_date"])
            db.add(SkuAnalysis(
                product_id=product_id,
                stat_date=stat_date,
                sku_id=r.get("sku_code", ""),
                sku_info=r.get("sku_info", ""),
                metrics=r["metrics"],
            ))
            imported += 1

    elif file_type == "service_data":
        for r in records:
            product_id = sku_map.get(r["sku_id"])
            if not product_id:
                skipped += 1
                continue
            stat_date = date.fromisoformat(r["stat_date"])
            db.add(ServiceData(
                product_id=product_id,
                stat_date=stat_date,
                metrics=r["metrics"],
            ))
            imported += 1

    await db.flush()
    return {"status": "ok", "imported": imported, "skipped": skipped, "total": len(records)}


def _create_traffic_source(db: AsyncSession, r: dict) -> None:
    """创建流量来源记录（无需 sku_id）。"""
    stat_date = date.fromisoformat(r["stat_date"])
    db.add(TrafficSource(
        product_id=1,  # 占位，流量来源是店铺级数据
        stat_date=stat_date,
        source_name=r["source_name"],
        sub_source=r.get("sub_source", ""),
        metrics=r["metrics"],
    ))


# ── 商品导入（保留原有逻辑）──────────────────────

@router.post("/import-from-export")
async def import_from_export(
    filename: str,
    default_cost_price: float = Query(..., description="统一进货成本价 (USD)"),
    _api_key: str = Depends(verify_api_key),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """将指定的导出文件中的商品导入系统。"""
    if default_cost_price <= 0:
        raise HTTPException(status_code=400, detail="成本价必须大于 0")

    files = _find_downloaded_xlsx_files()
    matched = [f for f in files if f["name"] == filename]
    if not matched:
        raise HTTPException(status_code=404, detail=f"文件 '{filename}' 未找到或已过期")

    file_type = matched[0]["file_type"]
    if file_type != "product_list":
        raise HTTPException(status_code=400, detail="该文件不是商品列表，请使用分析数据导入")

    file_path = matched[0]["path"]
    try:
        products = _parse_product_list(file_path)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"解析文件失败: {exc}")

    if not products:
        raise HTTPException(status_code=400, detail="文件中没有有效商品数据")

    existing_result = await db.execute(select(Product.sku_id))
    existing_sku_ids = {r[0] for r in existing_result.all()}

    imported = 0
    skipped = 0
    errors: list[str] = []

    for p in products:
        sku_id = p["sku_id"]
        name = p.get("name", "")
        if not sku_id or not name:
            errors.append(f"SKU {sku_id or '?'}: 缺少名称")
            skipped += 1
            continue
        if sku_id in existing_sku_ids:
            skipped += 1
            continue

        product = Product(
            sku_id=sku_id,
            name=name,
            cost_price=default_cost_price,
            category=p.get("category") or None,
            is_tracked=True,
        )
        db.add(product)
        try:
            await db.flush()
            imported += 1
        except IntegrityError:
            await db.rollback()
            skipped += 1

    await db.flush()
    return {
        "status": "ok",
        "imported": imported,
        "skipped": skipped,
        "total": len(products),
        "errors": errors[:10],
    }
