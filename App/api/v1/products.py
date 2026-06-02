"""Products CRUD — 商品成本录入与管理."""

from __future__ import annotations

import csv
import io
import re
from datetime import datetime, timezone
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, status
from sqlalchemy import select, update
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from App.core.database import get_db
from App.core.logging import get_logger
from App.core.security import verify_api_key
from App.models.base import Product
from App.schemas.product import (
    CSVImportResult,
    ExportRequest,
    ProductCreate,
    ProductRead,
    ProductToggleTracking,
    ProductUpdate,
)
from App.services.cache_service import get_cache, set_cache

logger = get_logger(__name__)

router = APIRouter(prefix="/products", tags=["products"])


# ── CRUD ────────────────────────────────────────

@router.get("/", response_model=list[ProductRead])
async def list_products(
    tracked: bool | None = Query(None, description="Filter: true=仅已跟踪, false=仅未跟踪"),
    _api_key: str = Depends(verify_api_key),
    db: AsyncSession = Depends(get_db),
) -> list[ProductRead]:
    """列出所有商品。可选 ?tracked=true/false 过滤（缓存 60 秒）。"""
    cache_key = f"products:list:{tracked}"
    cached = await get_cache(cache_key)
    if cached is not None:
        return [ProductRead(**r) for r in cached]

    stmt = select(Product).order_by(Product.created_at.desc())
    if tracked is not None:
        stmt = stmt.where(Product.is_tracked == tracked)
    result = await db.execute(stmt)
    products = result.scalars().all()
    serializable = [
        {
            "id": p.id,
            "sku_id": p.sku_id,
            "name": p.name,
            "cost_price": float(p.cost_price),
            "category": p.category,
            "is_tracked": p.is_tracked,
            "created_at": p.created_at.isoformat() if p.created_at else None,
        }
        for p in products
    ]
    await set_cache(cache_key, serializable, ttl=60)
    return [ProductRead.model_validate(p) for p in products]


@router.get("/{product_id}", response_model=ProductRead)
async def get_product(
    product_id: int,
    _api_key: str = Depends(verify_api_key),
    db: AsyncSession = Depends(get_db),
) -> ProductRead:
    """获取单个商品详情。"""
    product = await db.get(Product, product_id)
    if product is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="商品不存在")
    return ProductRead.model_validate(product)


@router.post("/", response_model=ProductRead, status_code=status.HTTP_201_CREATED)
async def create_product(
    body: ProductCreate,
    _api_key: str = Depends(verify_api_key),
    db: AsyncSession = Depends(get_db),
) -> ProductRead:
    """创建新商品。"""
    product = Product(**body.model_dump())
    db.add(product)
    try:
        await db.flush()
        await db.refresh(product)
    except IntegrityError:
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"SKU ID '{body.sku_id}' 已存在",
        )
    return ProductRead.model_validate(product)


@router.put("/{product_id}", response_model=ProductRead)
async def update_product(
    product_id: int,
    body: ProductUpdate,
    _api_key: str = Depends(verify_api_key),
    db: AsyncSession = Depends(get_db),
) -> ProductRead:
    """更新商品信息（仅更新提供的字段）。"""
    product = await db.get(Product, product_id)
    if product is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="商品不存在")

    update_data = body.model_dump(exclude_unset=True)
    if not update_data:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail="没有提供需要更新的字段"
        )

    for field, value in update_data.items():
        setattr(product, field, value)

    try:
        await db.flush()
        await db.refresh(product)
    except IntegrityError:
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="更新失败：SKU ID 可能重复",
        )
    return ProductRead.model_validate(product)


@router.delete("/{product_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_product(
    product_id: int,
    _api_key: str = Depends(verify_api_key),
    db: AsyncSession = Depends(get_db),
) -> None:
    """删除商品。"""
    product = await db.get(Product, product_id)
    if product is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="商品不存在")
    await db.delete(product)
    await db.flush()


# ── 跟踪管理 ────────────────────────────────────

@router.put("/{product_id}/toggle-tracked", response_model=ProductRead)
async def toggle_product_tracking(
    product_id: int,
    body: ProductToggleTracking,
    _api_key: str = Depends(verify_api_key),
    db: AsyncSession = Depends(get_db),
) -> ProductRead:
    """切换单个商品的跟踪状态。"""
    product = await db.get(Product, product_id)
    if product is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="商品不存在")
    product.is_tracked = body.is_tracked
    await db.flush()
    await db.refresh(product)
    return ProductRead.model_validate(product)


@router.post("/batch-set-tracking")
async def batch_set_tracking(
    tracked_ids: list[int],
    untracked_ids: list[int] | None = None,
    _api_key: str = Depends(verify_api_key),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """批量设置商品跟踪状态。tracked_ids 设为跟踪，untracked_ids 取消跟踪。"""
    tracked = 0
    untracked = 0
    if tracked_ids:
        await db.execute(
            update(Product).where(Product.id.in_(tracked_ids)).values(is_tracked=True)
        )
        tracked = len(tracked_ids)
    if untracked_ids:
        await db.execute(
            update(Product).where(Product.id.in_(untracked_ids)).values(is_tracked=False)
        )
        untracked = len(untracked_ids)
    await db.flush()
    return {"status": "ok", "tracked_count": tracked, "untracked_count": untracked}


@router.post("/batch-track")
async def batch_track(
    tracked_ids: list[int],
    _api_key: str = Depends(verify_api_key),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """批量设置商品为已跟踪（兼容旧路径）。"""
    if not tracked_ids:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="tracked_ids 不能为空")
    await db.execute(
        update(Product).where(Product.id.in_(tracked_ids)).values(is_tracked=True)
    )
    await db.flush()
    return {"status": "ok", "tracked_count": len(tracked_ids)}


# ── 批量导入 (CSV + XLSX) ──────────────────────

def _detect_delimiter(sample: str) -> str:
    """自动检测 CSV 分隔符（逗号或制表符）。"""
    first_line = sample.split("\n")[0] if sample else ""
    tabs = first_line.count("\t")
    commas = first_line.count(",")
    return "\t" if tabs > commas else ","


def _normalize_columns(field_names: list[str]) -> tuple[dict[str, str], set[str]]:
    """标准化列名：去空格、转小写、别名映射。返回 (field_map, mapped_fields)。"""
    ALIASES = {
        "sku": "sku_id", "商品id": "sku_id", "商品编号": "sku_id",
        "product": "sku_id", "product_id": "sku_id",
        "id": "sku_id",
        "商品名称": "name", "商品名": "name", "product_name": "name",
        "*商品标题": "name", "商品标题": "name",
        "成本": "cost_price", "成本价": "cost_price", "价格": "cost_price",
        "*货值(usd)": "cost_price", "*货值": "cost_price",
        "类目": "category", "分类": "category", "品类": "category",
        "*零售价(usd)": "price", "*零售价": "price",
    }
    field_map: dict[str, str] = {}
    for fn in field_names:
        key = fn.strip().lower().replace(" ", "_").replace("-", "_")
        if key in ALIASES:
            key = ALIASES[key]
        field_map[fn] = key
    return field_map, set(field_map.values())


def _parse_csv_rows(content: str) -> list[dict]:
    """解析 CSV 内容为 list[dict]。"""
    delimiter = _detect_delimiter(content)
    reader = csv.DictReader(io.StringIO(content), delimiter=delimiter)
    if reader.fieldnames is None:
        raise ValueError("文件没有表头")
    field_map, mapped_fields = _normalize_columns(reader.fieldnames)
    required = {"sku_id", "name", "cost_price"}
    missing = required - mapped_fields
    if missing:
        raise ValueError(
            f"缺少必填列: {', '.join(missing)}。\n"
            f"需要: sku_id, name, cost_price\n"
            f"当前表头: {', '.join(reader.fieldnames)}"
        )
    rows: list[dict] = []
    for row in reader:
        mapped: dict[str, str] = {}
        for original_key, value in row.items():
            mapped[field_map.get(original_key, original_key.strip().lower())] = (value or "").strip()
        rows.append(mapped)
    return rows


def _parse_xlsx_rows(file_bytes: bytes) -> list[dict]:
    """解析速卖通导出 XLSX 为 list[dict]。

    兼容两种格式：
    - 商品信息（类目+SPU级别）：id(→sku_id), *商品标题(→name)
    - SKU信息（规格级别）：id(→sku_id), *商品标题(→name), *零售价(USD)
    每个 sheet 为一个类目，自动去重（按商品 id）。

    注意：导出文件不包含成本价（cost_price），导入时前端会提示用户统一输入。
    """
    try:
        import openpyxl
    except ImportError:
        raise ValueError("不支持 XLSX 格式，请安装 openpyxl: pip install openpyxl")

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    ALL_ALIASES = {
        "id": "sku_id", "sku": "sku_id",
        "*商品标题": "name", "商品标题": "name", "商品名称": "name",
    }

    seen_sku = set()
    all_rows: list[dict] = []
    total_data_rows = 0

    for sname in wb.sheetnames:
        if "_hide" in sname:  # 跳过隐藏表
            continue
        ws = wb[sname]
        if ws.max_row < 2:
            continue

        # 读取表头（第2行）
        headers_raw: list[str] = []
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row=2, column=col).value
            headers_raw.append(str(v).strip() if v else "")

        # 标准化表头
        field_map: dict[int, str] = {}
        mapped_set: set[str] = set()
        for col_idx, raw in enumerate(headers_raw):
            key = raw.lower().replace(" ", "_").replace("-", "_")
            key = ALL_ALIASES.get(key, key)
            field_map[col_idx] = key
            mapped_set.add(key)

        has_cost = "cost_price" in mapped_set

        # 读取数据行（第3行起）
        for row_idx in range(3, ws.max_row + 1):
            # 检查是否全空行
            row_vals: list[str] = []
            has_any = False
            for col in range(1, ws.max_column + 1):
                v = ws.cell(row=row_idx, column=col).value
                s = ""
                if v is not None:
                    if isinstance(v, float):
                        s = str(int(v)) if v == int(v) else str(v)
                    else:
                        s = str(v).strip()
                    if s:
                        has_any = True
                row_vals.append(s)

            if not has_any:
                continue

            total_data_rows += 1

            # 构建行 dict
            row_dict: dict[str, str] = {}
            for col_idx, key in field_map.items():
                val = row_vals[col_idx] if col_idx < len(row_vals) else ""
                if key in row_dict:
                    # 同名列已存在，取非空值
                    if val and not row_dict[key]:
                        row_dict[key] = val
                else:
                    row_dict[key] = val

            sku_id = row_dict.get("sku_id", "")
            if not sku_id:
                continue

            # 去重：相同 sku_id 只保留第一次出现
            if sku_id in seen_sku:
                continue
            seen_sku.add(sku_id)

            # 填入类目
            if not row_dict.get("category"):
                row_dict["category"] = sname

            all_rows.append(row_dict)

    wb.close()

    if total_data_rows == 0:
        raise ValueError("文件中没有数据行")

    if not all_rows:
        raise ValueError(
            f"文件中没有找到有效商品数据（共 {total_data_rows} 行，"
            f"但缺少 sku_id 列或所有行均为空）"
        )

    return all_rows


@router.post("/import", response_model=CSVImportResult)
async def import_products(
    file: UploadFile,
    preview: bool = Query(False, description="预览模式：仅解析返回数据，不写入数据库"),
    default_cost_price: float | None = Query(None, description="XLSX 中无成本价时使用的默认值"),
    _api_key: str = Depends(verify_api_key),
    db: AsyncSession = Depends(get_db),
) -> CSVImportResult:
    """批量导入商品 — 支持 CSV / TSV / XLSX（含速卖通导出格式）。

    自动检测文件格式和分隔符，支持列名别名（id→sku_id, *商品标题→name 等）。
    XLSX 支持多 sheet 类目自动归类、SKU 级别去重。
    设置 ?preview=true 可预览解析结果而不写入数据库。
    """
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="文件为空")

    fname = file.filename or "unknown"
    ext = fname.lower().rsplit(".", 1)[-1] if "." in fname else ""

    if ext in ("xlsx", "xls"):
        try:
            rows = _parse_xlsx_rows(raw)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc))
    elif ext in ("csv", "tsv", "txt"):
        try:
            content = raw.decode("utf-8-sig")
        except UnicodeDecodeError:
            try:
                content = raw.decode("gbk")
            except UnicodeDecodeError:
                raise HTTPException(status_code=400, detail="文件编码无法识别，请使用 UTF-8 或 GBK")
        try:
            rows = _parse_csv_rows(content)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc))
    else:
        raise HTTPException(
            status_code=400,
            detail=f"不支持的文件格式 (.{ext})。支持: CSV, TSV, TXT, XLSX",
        )

    total_rows = len(rows)
    if total_rows == 0:
        raise HTTPException(status_code=400, detail="文件中没有数据行")

    # 预览模式：返回解析结果但不写入
    if preview:
        has_cost_col = any("cost_price" in r for r in rows) or any(
            r.get("cost_price", "").strip() for r in rows
        )
        preview_rows = [
            {
                "row": idx,
                "sku_id": r.get("sku_id", ""),
                "name": r.get("name", ""),
                "cost_price": r.get("cost_price", ""),
                "category": r.get("category", ""),
            }
            for idx, r in enumerate(rows, start=2)
        ]
        result = CSVImportResult(
            total_rows=total_rows,
            success_count=0,
            failed_rows=[],
            preview_rows=preview_rows,
        )
        # 标记是否缺少成本价（前端可据此显示默认价格输入框）
        if not has_cost_col:
            result.missing_cost_price = True
        return result

    success_count = 0
    failed_rows: list[dict] = []

    for idx, row in enumerate(rows, start=2):
        sku_id = (row.get("sku_id") or "").strip()
        name = (row.get("name") or "").strip()
        cost_price_raw = (row.get("cost_price") or "").strip()
        category = row.get("category") or None

        if not sku_id:
            failed_rows.append({"row": idx, "sku_id": "", "error": "sku_id 为空"})
            continue
        if not name:
            failed_rows.append({"row": idx, "sku_id": sku_id, "error": "name 为空"})
            continue

        # 如果行内没有成本价，使用默认值
        if not cost_price_raw:
            if default_cost_price is not None:
                cost_price = default_cost_price
            else:
                failed_rows.append({"row": idx, "sku_id": sku_id, "error": "缺少成本价，请在导入时指定默认成本价"})
                continue
        else:
            cleaned = re.sub(r'[¥$€\s,]', '', cost_price_raw)
            try:
                cost_price = float(cleaned)
            except (ValueError, TypeError):
                failed_rows.append({"row": idx, "sku_id": sku_id, "error": f"成本价无效: {cost_price_raw}"})
                continue
        if cost_price <= 0:
            failed_rows.append({"row": idx, "sku_id": sku_id, "error": "cost_price 必须大于 0"})
            continue

        existing = await db.execute(select(Product).where(Product.sku_id == sku_id))
        if existing.scalar_one_or_none() is not None:
            failed_rows.append({"row": idx, "sku_id": sku_id, "error": "SKU ID 已存在"})
            continue

        db.add(Product(sku_id=sku_id, name=name, cost_price=cost_price, category=category))
        try:
            await db.flush()
            success_count += 1
        except IntegrityError:
            await db.rollback()
            failed_rows.append({"row": idx, "sku_id": sku_id, "error": "数据库写入冲突"})

    return CSVImportResult(total_rows=total_rows, success_count=success_count, failed_rows=failed_rows)


@router.post("/import-csv", response_model=CSVImportResult)
async def import_csv_legacy(
    file: UploadFile,
    _api_key: str = Depends(verify_api_key),
    db: AsyncSession = Depends(get_db),
) -> CSVImportResult:
    """(旧接口) 通过 CSV 文件批量导入商品 — 重定向到新版 /import。"""
    return await import_products(file=file, _api_key=_api_key, db=db)


# ── 导出 ────────────────────────────────────────

@router.post("/export")
async def export_products(
    body: ExportRequest | None = None,
    _api_key: str = Depends(verify_api_key),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """导出商品为 CSV。支持中文表头，可直接在 Excel 中打开。

    如果提供 sku_ids 则仅导出选中商品，否则导出全部。
    """
    sku_ids = body.sku_ids if body else None
    stmt = select(Product).order_by(Product.created_at.desc())
    if sku_ids:
        stmt = stmt.where(Product.sku_id.in_(sku_ids))
    result = await db.execute(stmt)
    products = result.scalars().all()

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["SKU ID", "商品名称", "成本价(USD)", "类目", "是否跟踪", "创建时间"])
    for p in products:
        writer.writerow([
            p.sku_id,
            p.name,
            float(p.cost_price),
            p.category or "",
            "是" if p.is_tracked else "否",
            p.created_at.strftime("%Y-%m-%d %H:%M:%S") if p.created_at else "",
        ])

    return {
        "status": "ok",
        "filename": f"商品导出_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.csv",
        "content": buf.getvalue(),
        "count": len(products),
    }
