"""单品分析历史数据采集 — Playwright + API 拦截方案.

采集流程（串行逐类型，每个类型独立 try/except）：
  核心指标 → 流量 → 关键词 → 服务 → SKU

每个类型完成立即写入数据库（不等全部完成）。
如果单次最大日期范围 < 请求的总范围，自动拆分时间段循环采集。

设计原则：
  - 不注入 stealth.js（SYCM SPA 会因此不加载）
  - 优先通过 XHR API 响应获取数据（mtop.table.query），而非导出文件
  - 日期设置通过 JS 直接操控 ant-design DatePicker 组件
  - 每个 Tab 的日期选择器完全独立操作，不共享状态
"""

from __future__ import annotations

import json
import random
import time
from datetime import date, datetime, timedelta, timezone
from typing import TYPE_CHECKING, Any

if TYPE_CHECKING:
    from sqlalchemy.ext.asyncio import AsyncSession
    from App.services.cookie_manager import CookieManager

from App.core.logging import get_logger

logger = get_logger(__name__)

# ── 页面 URL ───────────────────────────────────────
SYCM_SEARCH_URL = "https://csp.aliexpress.com/apps/sycm/product/search"
DETAIL_URL_TEMPLATE = (
    "https://csp.aliexpress.com/m_apps/sycm/product-analyse"
    "?productId={product_id}&channelId=363432"
)

# ── 各 Tab 配置 ────────────────────────────────────
TAB_CONFIGS: dict[str, dict[str, Any]] = {
    "core": {
        "name": "核心指标",
        "max_days": 180,
        "date_type": "single",
        "date_placeholder": "请选择日期",
        "export_btn": "导出",
        "model_name": "CoreMetric",
    },
    "traffic": {
        "name": "流量",
        "max_days": 30,
        "date_type": "single",
        "date_placeholder": "请选择日期",
        "export_btn": "明细数据下载",
        "model_name": "TrafficSource",
    },
    "keyword": {
        "name": "关键词",
        "max_days": 90,
        "date_type": "range",
        "date_start_placeholder": "起始日期",
        "date_end_placeholder": "结束日期",
        "export_btn": "导出",
        "model_name": "KeywordData",
    },
    "service": {
        "name": "服务",
        "max_days": 90,
        "date_type": "single",
        "date_placeholder": "请选择日期",
        "export_btn": "导出",
        "model_name": "ServiceData",
    },
    "sku": {
        "name": "SKU",
        "max_days": 90,
        "date_type": "single",
        "date_placeholder": "请选择日期",
        "export_btn": "导出",
        "model_name": "SkuAnalysis",
    },
}

COLLECTION_ORDER = ["core", "traffic", "keyword", "service", "sku"]

# 数据 API URL 特征
DATA_API_PATTERN = "mtop.aliexpress.seller.business.advice.table.query"
PRICE_API_PATTERNS = ["mtop.aliexpress.dps.query"]


# ── 日期分段工具 ──────────────────────────────────
def _split_date_range(date_from: date, date_to: date, max_days: int) -> list[tuple[date, date]]:
    """将日期范围按 max_days 拆分为多个时间段，从 date_to 向前滚动。"""
    chunks: list[tuple[date, date]] = []
    current_end = date_to
    while current_end > date_from:
        chunk_start = max(date_from, current_end - timedelta(days=max_days - 1))
        chunks.append((chunk_start, current_end))
        current_end = chunk_start - timedelta(days=1)
    return list(reversed(chunks))


# ── 防检测辅助 ─────────────────────────────────────
def _random_delay(min_ms: int = 300, max_ms: int = 1200) -> float:
    return random.randint(min_ms, max_ms) / 1000


# ── 同步采集主函数 ────────────────────────────────
def _run_analytics_sync(
    product_id: str,
    date_from: date,
    date_to: date,
    cookies: list[dict],
    headless: bool = False,
) -> dict:
    """在同步线程中执行全部 5 类数据采集。返回结构化结果。"""
    from App.services.browser import BrowserService

    result: dict = {
        "success": True,
        "product_id": product_id,
        "date_from": date_from.isoformat(),
        "date_to": date_to.isoformat(),
        "collected_at": datetime.now(timezone.utc).isoformat(),
        "tabs": {},
        "errors": [],
        "duration_seconds": 0,
    }

    t0 = time.perf_counter()
    browser_svc: BrowserService | None = None

    try:
        browser_svc = BrowserService(headless=headless)
        context = browser_svc.new_context(cookies=cookies)
        page = context.new_page()

        # ── 1. 导航到搜索页 → 搜索 → 进入详情 ──
        if not _navigate_to_detail(page, product_id):
            result["success"] = False
            result["errors"].append("无法进入单品分析详情页")
            return result

        # ── 2. 逐个 Tab 采集 ──
        for tab_type in COLLECTION_ORDER:
            config = TAB_CONFIGS[tab_type]
            tab_result = _collect_tab(page, tab_type, config, product_id, date_from, date_to)
            result["tabs"][tab_type] = tab_result
            if tab_result.get("error"):
                result["errors"].append(f"[{tab_type}] {tab_result['error']}")
                logger.warning("Tab %s 采集失败: %s", tab_type, tab_result["error"])

        page.close()
        context.close()

    except Exception as exc:
        logger.exception("采集流程异常")
        result["success"] = False
        result["errors"].append(f"采集流程异常: {exc}")
    finally:
        if browser_svc is not None:
            browser_svc.close()
        result["duration_seconds"] = round(time.perf_counter() - t0, 2)

    return result


def _navigate_to_detail(page, product_id: str) -> bool:
    """导航到搜索页 → 搜索商品 ID → 进入详情页。返回是否成功。"""
    page.goto(SYCM_SEARCH_URL, wait_until="domcontentloaded", timeout=60000)
    time.sleep(_random_delay(5000, 8000))

    if "login" in page.url.lower():
        logger.error("导航后重定向到登录页，Cookie 可能已过期")
        return False

    # 搜索框
    try:
        si = page.wait_for_selector("input[name='inputItemId']", timeout=30000)
        si.click()
        si.fill("")
        si.type(product_id, delay=30)
        time.sleep(_random_delay(300, 800))

        sb = page.query_selector("button:has-text('搜索')")
        if sb:
            sb.click()
        else:
            page.keyboard.press("Enter")
    except Exception as e:
        logger.error("搜索输入失败: %s", e)
        return False

    time.sleep(_random_delay(3000, 5000))

    # 检查是否已跳转到详情页
    if "product-analyse" in page.url:
        return True

    # 尝试点击第一个 "查询单品" 按钮
    try:
        query_btn = page.query_selector("button:has-text('查询单品')")
        if query_btn:
            query_btn.click()
            time.sleep(_random_delay(3000, 5000))
            if "product-analyse" in page.url:
                return True
    except Exception:
        pass

    logger.warning("未能进入详情页，当前 URL: %s", page.url)
    return False


def _collect_tab(
    page, tab_type: str, config: dict, product_id: str,
    date_from: date, date_to: date,
) -> dict:
    """采集单个 Tab 的数据。按时间段分片循环。"""
    tab_result: dict = {
        "tab_type": tab_type,
        "records": [],
        "chunks": 0,
        "error": None,
    }

    try:
        if not _click_tab(page, config["name"]):
            tab_result["error"] = f"无法点击 Tab '{config['name']}'"
            return tab_result

        time.sleep(_random_delay(3000, 5000))

        # 拆分时间段
        chunks = _split_date_range(date_from, date_to, config["max_days"])
        tab_result["chunks"] = len(chunks)

        for chunk_start, chunk_end in chunks:
            logger.info("  %s 采集时间段: %s ~ %s", config["name"], chunk_start, chunk_end)

            # 设置日期
            _set_date_on_tab(page, config, chunk_start, chunk_end)

            # 等待 API 响应并提取数据
            records = _capture_tab_data(page, tab_type, product_id, chunk_start, chunk_end)

            # 如果 API 拦截未获取到数据，尝试导出按钮
            if not records and config.get("export_btn"):
                logger.info("  API 无数据，尝试导出 '%s'", config["export_btn"])
                export_records = _try_export_download(page, config, product_id)
                if export_records:
                    records = export_records
                    logger.info("  导出成功: %d 条", len(records))

            if records:
                tab_result["records"].extend(records)
                logger.info("  %s 时间段 %s~%s 采集到 %d 条", config["name"], chunk_start, chunk_end, len(records))
            else:
                logger.warning("  %s 时间段 %s~%s 未采集到数据", config["name"], chunk_start, chunk_end)

    except Exception as e:
        logger.exception("采集 Tab %s 异常", config["name"])
        tab_result["error"] = str(e)

    return tab_result


def _click_tab(page, tab_name: str) -> bool:
    """点击指定名称的 Tab。"""
    # 方式1: Playwright text selector
    try:
        el = page.query_selector(f"text='{tab_name}'")
        if el and el.is_visible():
            el.click()
            return True
    except Exception:
        pass

    # 方式2: JS TreeWalker (无视 DOM 层级)
    try:
        return page.evaluate(f"""() => {{
            const w = document.createTreeWalker(document.body, 4, null, false);
            let n;
            while (n = w.nextNode()) {{
                if (n.textContent.trim() === '{tab_name}' && n.offsetParent !== null) {{
                    n.click();
                    return true;
                }}
            }}
            return false;
        }}""")
    except Exception:
        return False


def _set_date_on_tab(page, config: dict, chunk_start: date, chunk_end: date) -> None:
    """通过 JS 设置 ant-design DatePicker 的值并触发 change 事件。"""
    if config["date_type"] == "range":
        _set_antd_date(page, config["date_start_placeholder"], chunk_start)
        time.sleep(_random_delay(200, 500))
        _set_antd_date(page, config["date_end_placeholder"], chunk_end)
    else:
        _set_antd_date(page, config["date_placeholder"], chunk_end)

    # 设置后等待页面重新加载数据
    time.sleep(_random_delay(2000, 4000))


def _set_antd_date(page, placeholder: str, target_date: date) -> None:
    """通过原生 value setter 触发 React controlled component 的 onChange。

    ant-design 的 DatePicker 是 controlled 组件，监听 input/change 事件。
    使用 Object.getOwnPropertyDescriptor 绕过 React 的 value 覆盖。
    """
    date_str = target_date.strftime("%Y-%m-%d")
    page.evaluate(f"""() => {{
        const input = document.querySelector('input[placeholder="{placeholder}"]');
        if (!input) return;
        const nativeSetter = Object.getOwnPropertyDescriptor(
            window.HTMLInputElement.prototype, 'value'
        ).set;
        nativeSetter.call(input, '{date_str}');
        input.dispatchEvent(new Event('change', {{ bubbles: true }}));
    }}""")


def _try_export_download(page, config: dict, product_id: str) -> list[dict]:
    """备选方案：点击导出按钮 → 等待下载 xlsx → 解析为记录。"""
    import tempfile
    import os

    temp_dir = tempfile.mkdtemp(prefix="ae_export_")
    download_paths: list[str] = []

    def _on_download(download):
        ext = os.path.splitext(download.suggested_filename)[1].lower()
        if ext in (".xlsx", ".xls"):
            dest = os.path.join(temp_dir, download.suggested_filename)
            try:
                download.save_as(dest)
                download_paths.append(dest)
            except Exception:
                pass

    page.on("download", _on_download)

    export_btn_text = config["export_btn"]
    try:
        btn = page.query_selector(f"button:has-text('{export_btn_text}')")
        if btn and btn.is_visible():
            btn.click()
        else:
            target = "明细数据下载" if export_btn_text == "明细数据下载" else "导出"
            btn = page.query_selector(f"text='{target}'")
            if btn and btn.is_visible():
                btn.click()
            else:
                logger.warning("  导出按钮 '%s' 未找到", export_btn_text)
                page.remove_listener("download", _on_download)
                return []
    except Exception as e:
        logger.warning("  点击导出按钮失败: %s", e)
        page.remove_listener("download", _on_download)
        return []

    deadline = time.monotonic() + 45
    while time.monotonic() < deadline and len(download_paths) == 0:
        time.sleep(1)

    page.remove_listener("download", _on_download)

    if not download_paths:
        logger.warning("  导出超时（45秒），未收到文件")
        _cleanup_temp(temp_dir)
        return []

    records = _parse_export_xlsx(download_paths[0])
    _cleanup_temp(temp_dir)
    return records


def _cleanup_temp(temp_dir: str) -> None:
    """清理临时目录。"""
    try:
        import shutil
        shutil.rmtree(temp_dir, ignore_errors=True)
    except Exception:
        pass


def _parse_export_xlsx(file_path: str) -> list[dict]:
    """解析导出的 xlsx 文件为 dict 列表。"""
    import openpyxl

    records: list[dict] = []
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        logger.warning("  无法打开导出文件: %s", e)
        return []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        header_row = list(rows[1])
        headers = [str(h).strip() if h is not None else "" for h in header_row]

        if not any(h for h in headers):
            continue

        for row in rows[2:]:
            if not row:
                continue
            rec: dict = {}
            for i, val in enumerate(row):
                if i < len(headers) and headers[i]:
                    if val is None:
                        continue
                    if isinstance(val, (int, float)):
                        rec[headers[i]] = val
                    else:
                        s = str(val).strip()
                        if s and s != "None":
                            rec[headers[i]] = s
            if any(v for v in rec.values()):
                records.append(rec)

    wb.close()
    return records


def _capture_tab_data(
    page, tab_type: str, product_id: str,
    chunk_start: date, chunk_end: date,
) -> list[dict]:
    """捕获当前 Tab 的数据 API 响应并提取结构化的记录。"""
    captured: list[dict] = []

    def _on_response(response):
        url = response.url
        if DATA_API_PATTERN not in url and not any(p in url for p in PRICE_API_PATTERNS):
            return
        try:
            body = response.json()
            extracted = _extract_table_rows(body)
            if extracted:
                captured.extend(extracted)
        except Exception:
            pass

    page.on("response", _on_response)

    # 轮询等待最多 20 秒获取 API 响应
    deadline = time.monotonic() + 20
    polls = 0
    while time.monotonic() < deadline:
        time.sleep(1)
        polls += 1
        # 如果已有数据，提前退出
        if captured:
            logger.debug("API 数据已捕获（%d 条），耗时 %d 秒", len(captured), polls)
            break
        # 每 5 秒额外检查一次页面是否还在正常状态
        if polls % 5 == 0 and "login" in page.url.lower():
            logger.warning("采集过程中跳转到登录页")
            break

    page.remove_listener("response", _on_response)

    # 为每条记录标记 tab_type 和日期
    for rec in captured:
        rec["_tab_type"] = tab_type
        rec["_stat_date"] = chunk_end.isoformat()

    return captured


# ── JSON 数据提取 ──────────────────────────────────
def _extract_table_rows(obj: Any, depth: int = 0) -> list[dict]:
    """递归搜索 JSON 响应，提取表格数据行。

    速卖通 mtop API 返回格式：
      {"api":"...","data":{"data":{"table":{"dataSource":[{"col1":"val","col2":"val"},...]}}}}
    或
      {"data":{"dataList":[...]}}
    或
      {"data":{"result":{"data":[...]}}}
    """
    if depth > 8:
        return []

    if isinstance(obj, dict):
        # mtop 格式
        if "api" in obj and "data" in obj and "ret" in obj:
            data_wrapper = obj.get("data")
            if isinstance(data_wrapper, dict):
                inner = data_wrapper.get("data")
                if isinstance(inner, dict):
                    # 表格数据: {"table":{"dataSource":[...]}}
                    table = inner.get("table")
                    if isinstance(table, dict):
                        ds = table.get("dataSource")
                        if isinstance(ds, list) and ds:
                            return [
                                {k: v for k, v in row.items() if v is not None}
                                for row in ds
                            ]
                    for list_key in ("dataList", "list", "records", "items"):
                        candidate = inner.get(list_key)
                        if isinstance(candidate, list) and candidate:
                            return _extract_rows_from_list(candidate)
                model = data_wrapper.get("model")
                if isinstance(model, dict):
                    result_list = model.get("result", model.get("list", model.get("data")))
                    if isinstance(result_list, list) and result_list:
                        if isinstance(result_list[0], dict):
                            return result_list
                result_obj = data_wrapper.get("result")
                if isinstance(result_obj, dict):
                    for k in ("data", "list", "items", "records"):
                        v = result_obj.get(k)
                        if isinstance(v, list) and v:
                            return _extract_rows_from_list(v)
            return []

        # 通用 dict 递归
        for key, value in obj.items():
            if key.lower() in ("datalist", "records", "items", "list", "rows", "content", "data"):
                if isinstance(value, list) and value:
                    return _extract_rows_from_list(value)
            elif isinstance(value, (dict, list)):
                result = _extract_table_rows(value, depth + 1)
                if result:
                    return result

    elif isinstance(obj, list):
        return _extract_rows_from_list(obj)

    return []


def _extract_rows_from_list(lst: list) -> list[dict]:
    """从列表中提取 dict 行。"""
    result = []
    for item in lst:
        if isinstance(item, dict):
            result.append({k: v for k, v in item.items() if v is not None})
        elif isinstance(item, list):
            # 可能每行是数组，尝试猜测列名
            pass
    return result


# ── CSP 导出公共 API（供 data_collector 调用） ─────

# 列名映射：SYCM 核心指标导出中文列名 -> AdSnapshot 字段
CORE_EXPORT_COLUMN_MAP: dict[str, str] = {
    "展现量": "impressions",
    "展现": "impressions",
    "曝光量": "impressions",
    "曝光": "impressions",
    "点击量": "clicks",
    "点击": "clicks",
    "点击率": "ctr",
    "订单量": "orders",
    "订单": "orders",
    "转化率": "conversion_rate",
    "花费": "ad_spend",
    "消耗": "ad_spend",
    "广告花费": "ad_spend",
    "商品点击花费": "ad_spend",
    "推广花费": "ad_spend",
    "销售额": "revenue",
    "收入": "revenue",
    "成交金额": "revenue",
    "商品ID": "sku_id",
    "商品id": "sku_id",
    "productId": "sku_id",
    "product_id": "sku_id",
    "日期": "stat_date",
    "广告类型": "ad_type",
    "推广类型": "ad_type",
}

TRAFFIC_EXPORT_COLUMN_MAP: dict[str, str] = {
    "来源": "source_name",
    "来源名称": "source_name",
    "访客数": "visitors",
    "浏览量": "page_views",
    "点击量": "clicks",
    "花费": "ad_spend",
    "广告花费": "ad_spend",
    "推广花费": "ad_spend",
    "订单量": "orders",
    "转化率": "conversion_rate",
    "成交金额": "revenue",
    "销售额": "revenue",
    "商品点击花费": "ad_spend",
}


def export_product_ad_data_sync(
    page,
    product_id: str,
) -> list[dict]:
    """导航到 SYCM 单品分析详情页 → 导出核心指标数据 → 解析 XLSX 为记录。

    Returns:
        list[dict]: 每行一条记录，key 为中文列名，value 为各列数据。
    """
    if not _navigate_to_detail(page, product_id):
        logger.warning("export_product_ad_data_sync: 无法进入单品分析详情页 (product_id=%s)", product_id)
        return []

    # 点击核心指标 Tab
    if not _click_tab(page, "核心指标"):
        logger.warning("export_product_ad_data_sync: 无法点击核心指标 Tab")
        return []

    time.sleep(_random_delay(2000, 4000))

    config = TAB_CONFIGS["core"]
    records = _try_export_download(page, config, product_id)
    if not records:
        # 回退：尝试从流量来源 Tab 获取含花费的数据
        logger.info("核心指标无导出数据，尝试流量来源 Tab")
        time.sleep(_random_delay(1000, 2000))
        if _click_tab(page, "流量"):
            time.sleep(_random_delay(2000, 4000))
            traffic_config = TAB_CONFIGS["traffic"]
            records = _try_export_download(page, traffic_config, product_id)

    return records


def map_export_records_to_ad_snapshot(
    records: list[dict],
    sku_id: str,
    column_map: dict[str, str] | None = None,
) -> list[dict]:
    """将 SYCM 导出的原始记录映射为 AdSnapshot 兼容的 dict。

    Args:
        records: export_product_ad_data_sync 返回的原始记录列表。
        sku_id: 当前 SKU 的商品 ID。
        column_map: 列名映射字典，默认使用 CORE_EXPORT_COLUMN_MAP。

    Returns:
        list[dict]: 每条包含 AdSnapshot 字段（impressions, clicks, ctr 等）。
    """
    if not records:
        return []

    if column_map is None:
        column_map = CORE_EXPORT_COLUMN_MAP

    results: list[dict] = []
    for row in records:
        snapshot: dict = {
            "sku_id": sku_id,
            "impressions": 0,
            "clicks": 0,
            "ctr": 0.0,
            "orders": 0,
            "conversion_rate": 0.0,
            "ad_spend": 0.0,
            "revenue": 0.0,
            "ad_type": "standard",
        }

        for chinese_col, value in row.items():
            chinese_clean = chinese_col.strip()
            field = column_map.get(chinese_clean)
            if field is None:
                continue

            if field in ("impressions", "clicks", "orders"):
                try:
                    snapshot[field] = int(float(str(value).replace(",", "")))
                except (ValueError, TypeError):
                    pass
            elif field in ("ctr", "conversion_rate"):
                # 处理百分比格式: "12.34%" -> 0.1234
                try:
                    raw = str(value).replace("%", "").replace(",", "").strip()
                    val = float(raw)
                    if "百分比" in chinese_clean or "%" in str(value) or val > 1:
                        val = val / 100.0
                    snapshot[field] = val
                except (ValueError, TypeError):
                    pass
            elif field in ("ad_spend", "revenue"):
                try:
                    snapshot[field] = float(str(value).replace(",", "").replace("$", ""))
                except (ValueError, TypeError):
                    pass
            elif field in ("stat_date",):
                if value:
                    snapshot["stat_date"] = str(value).strip()

        # 至少要有一些关键字段才算有效
        if snapshot["impressions"] > 0 or snapshot["clicks"] > 0 or snapshot["ad_spend"] > 0:
            results.append(snapshot)

    return results


# ── 异步入口 ──────────────────────────────────────

async def collect_product_analytics(
    product_id: str,
    date_from: date,
    date_to: date,
    cookie_manager: CookieManager,
    db: AsyncSession,
    headless: bool = False,
) -> dict:
    """异步入口：执行一次完整的历史数据采集并逐类型写库。"""
    import asyncio
    from App.core.errors import ErrorCode, error_response
    from App.models.system_state import is_global_stop_active

    # ── 前置检查 ──
    cookies = await cookie_manager.load_cookies("aliexpress.com")
    if not cookies:
        return error_response(ErrorCode.COOKIE_MISSING, details={"action": "请先执行首次登录"})

    if await is_global_stop_active(db):
        return error_response(ErrorCode.GLOBAL_STOP, details={"action": "请检查警报中心并清除全局停止"})

    # ── 后台线程执行同步浏览器操作 ──
    loop = asyncio.get_event_loop()
    raw = await loop.run_in_executor(
        None, _run_analytics_sync, product_id, date_from, date_to, cookies, headless
    )

    # ── 逐类型写库 ──
    summary = {
        "success": raw.get("success", False),
        "product_id": product_id,
        "date_from": date_from.isoformat(),
        "date_to": date_to.isoformat(),
        "duration_seconds": raw.get("duration_seconds", 0),
        "tabs": {},
        "errors": raw.get("errors", []),
    }

    tab_model_map = {
        "core": ("CoreMetric", "core_metrics"),
        "traffic": ("TrafficSource", "traffic_sources"),
        "keyword": ("KeywordData", "keyword_data"),
        "service": ("ServiceData", "service_data"),
        "sku": ("SkuAnalysis", "sku_analyses"),
    }

    for tab_type in COLLECTION_ORDER:
        tab_result = raw.get("tabs", {}).get(tab_type, {})
        records = tab_result.get("records", [])
        if not records:
            summary["tabs"][tab_type] = {"saved": 0, "error": tab_result.get("error")}
            continue

        saved = await _write_tab_to_db(tab_type, records, product_id, db)
        summary["tabs"][tab_type] = {
            "saved": saved,
            "chunks": tab_result.get("chunks", 0),
        }
        if saved > 0:
            logger.info("写库 %s: %d 条", tab_type, saved)

    return summary


async def _write_tab_to_db(tab_type: str, records: list[dict], product_id: str, db: AsyncSession) -> int:
    """将单个 Tab 的采集结果写入数据库。"""
    from App.models.product_analytics import (
        CoreMetric, KeywordData, PriceDistribution, ServiceData, SkuAnalysis, TrafficSource,
    )

    model_map = {
        "core": CoreMetric,
        "traffic": TrafficSource,
        "keyword": KeywordData,
        "service": ServiceData,
        "sku": SkuAnalysis,
    }

    model_cls = model_map.get(tab_type)
    if not model_cls:
        return 0

    saved = 0
    now = datetime.now(timezone.utc)

    for rec in records:
        stat_date_str = rec.pop("_stat_date", None) or rec.pop("stat_date", None)
        if not stat_date_str:
            continue
        try:
            stat_date = datetime.strptime(stat_date_str[:10], "%Y-%m-%d").date()
        except (ValueError, TypeError):
            continue

        try:
            pid = int(product_id) if product_id.isdigit() else rec.get("product_id", 0)
        except (ValueError, TypeError):
            pid = 0
        if not pid:
            continue

        # 构建模型实例
        if tab_type == "traffic":
            instance = TrafficSource(
                product_id=pid,
                stat_date=stat_date,
                source_name=str(rec.get("来源", rec.get("sourceName", rec.get("source_name", "")))),
                sub_source=str(rec.get("二级来源", rec.get("subSource", rec.get("sub_source", ""))) or None),
                metrics=rec,
            )
        elif tab_type == "keyword":
            instance = KeywordData(
                product_id=pid,
                stat_date=stat_date,
                keyword=str(rec.get("关键词", rec.get("keyword", ""))),
                language=str(rec.get("语种", rec.get("language", "")) or None),
                metrics=rec,
            )
        elif tab_type == "sku":
            instance = SkuAnalysis(
                product_id=pid,
                stat_date=stat_date,
                sku_id=str(rec.get("skuId", rec.get("sku_id", rec.get("sku信息", "")))),
                sku_info=str(rec.get("sku信息", rec.get("skuInfo", rec.get("sku_info", ""))) or None),
                metrics=rec,
            )
        elif tab_type == "service":
            instance = ServiceData(
                product_id=pid,
                stat_date=stat_date,
                metrics=rec,
            )
        else:
            # core
            instance = CoreMetric(
                product_id=pid,
                stat_date=stat_date,
                metrics=rec,
            )

        db.add(instance)
        saved += 1

        # 每 50 条 flush 一次
        if saved % 50 == 0:
            await db.flush()

    if saved > 0:
        await db.flush()

    return saved
